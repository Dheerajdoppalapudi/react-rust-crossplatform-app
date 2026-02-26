import json
import os
import re

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from services.llm_service import default_llm_service
from .constants import THEMES, NUMBER_FORMATS, ALIGNMENTS

# Configuration
MAX_COLS = 100          # safety cap
MAX_FORMAT_ROWS = 10000  # cap for styling loops (data is never lost)
ENTERPRISE_FONT = "Calibri"


# Load system prompt from markdown file
_PROMPT_FILE = os.path.join(os.path.dirname(__file__), "excel_enhance_prompt.md")
with open(_PROMPT_FILE) as _f:
    SYSTEM_PROMPT = _f.read()

USER_PROMPT = """\
Workbook structure:

{structure}

Generate the formatting plan."""


# ---------------------------------------------------------------------------
# Sheet inspection helpers
# ---------------------------------------------------------------------------

def _detect_sheet_type(ws) -> str:
    """Classify a worksheet as 'hidden', 'chart', 'sparse', or 'data'."""
    if ws.sheet_state == "hidden":
        return "hidden"
    if ws.max_row is None or ws.max_column is None:
        return "chart"
    non_empty = sum(
        1
        for r in range(1, min(ws.max_row + 1, 21))
        for c in range(1, min(ws.max_column + 1, 21))
        if ws.cell(row=r, column=c).value is not None
    )
    total = min(ws.max_row, 20) * min(ws.max_column, 20)
    if total > 0 and non_empty / total < 0.05:
        return "sparse"
    return "data"


def _read_tab_color(ws) -> str | None:
    """Return existing tab color as a 6-char RRGGBB hex string, or None if unset."""
    tc = ws.sheet_properties.tabColor
    if tc is None:
        return None
    rgb = getattr(tc, "rgb", None)
    # "00000000" is the default/unset sentinel in openpyxl
    if rgb and rgb not in ("00000000", "FF000000"):
        return rgb[-6:]
    return None


def _detect_header_row(ws, max_col: int) -> tuple[int, bool]:
    """
    Return (header_row, has_title_row).

    A title row is row 1 when it contains only a single non-empty cell
    (or is merged across most of the sheet width) while row 2 holds
    the actual column headers.
    """
    row1_non_empty = sum(
        1 for c in range(1, max_col + 1)
        if ws.cell(row=1, column=c).value is not None
    )
    row2_non_empty = (
        sum(
            1 for c in range(1, max_col + 1)
            if ws.cell(row=2, column=c).value is not None
        )
        if ws.max_row and ws.max_row >= 2
        else 0
    )

    # Single populated cell in row 1 + multiple populated cells in row 2
    if row1_non_empty == 1 and row2_non_empty >= 2:
        return 2, True

    # Row 1 is a wide merge spanning at least half the data columns
    merged_title = any(
        m.min_row == 1
        and m.max_row == 1
        and (m.max_col - m.min_col) >= max(max_col // 2, 1)
        for m in ws.merged_cells.ranges
    )
    if merged_title and row2_non_empty >= 2:
        return 2, True

    return 1, False


# ===================================================================
# STEP 1 — Extract workbook structure (sent to LLM)
# ===================================================================
def extract_structure(filepath: str) -> list[dict]:
    """
    Read the workbook and return a lightweight summary per sheet.
    Hidden and chart-only sheets are excluded — they will not be formatted.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = []

    for name in wb.sheetnames:
        ws = wb[name]

        sheet_type = _detect_sheet_type(ws)
        existing_tab_color = _read_tab_color(ws)

        # Hidden / chart-only sheets carry no cell data — skip entirely
        if sheet_type in ("hidden", "chart"):
            continue

        if ws.max_row is None or ws.max_column is None or ws.max_row < 1:
            continue

        max_col = min(ws.max_column, MAX_COLS)
        total_rows = ws.max_row

        # --- Detect title row and actual header row ---
        header_row, has_title_row = _detect_header_row(ws, max_col)

        # --- Named Excel Tables (ranges defined by the workbook author) ---
        named_tables = []
        try:
            for tbl in ws.tables.values():
                named_tables.append({"name": tbl.name, "range": tbl.ref})
        except Exception:
            pass

        # --- Column headers (from detected header row) ---
        headers = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=header_row, column=c).value
            headers.append(str(val) if val is not None else "")

        # --- Sample data rows (up to 5 rows after the header) ---
        sample_rows = []
        for r in range(header_row + 1, min(total_rows + 1, header_row + 6)):
            row_vals = []
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                row_vals.append(str(val) if val is not None else "")
            sample_rows.append(row_vals)

        # --- Per-column metadata ---
        columns = []
        for ci in range(max_col):
            samples = [row[ci] for row in sample_rows if row[ci]]
            columns.append({
                "index": ci + 1,
                "letter": get_column_letter(ci + 1),
                "header": headers[ci],
                "sample_values": samples[:4],
                "inferred_type": _infer_type(samples),
                "non_empty_samples": len(samples),
            })

        # --- Merged cells ---
        merged = [str(m) for m in ws.merged_cells.ranges][:30]

        # --- Detect totals row (last row of the sheet) ---
        has_totals_row = False
        if total_rows > header_row + 1:
            first_cell = str(ws.cell(row=total_rows, column=1).value or "").lower()
            if first_cell in ("total", "totals", "grand total", "sum", "subtotal"):
                has_totals_row = True

        sheets.append({
            "sheet_name": name,
            "sheet_type": sheet_type,           # "data" or "sparse"
            "existing_tab_color": existing_tab_color,
            "has_title_row": has_title_row,
            "header_row": header_row,
            "named_tables": named_tables,
            "total_rows": total_rows,
            "total_columns": max_col,
            "has_totals_row": has_totals_row,
            "merged_cells": merged,
            "columns": columns,
        })

    wb.close()
    return sheets


def _infer_type(values: list[str]) -> str:
    """Heuristic type inference from sample string values."""
    if not values:
        return "text"

    scores = {"number": 0, "decimal": 0, "date": 0, "percentage": 0, "currency": 0, "text": 0}

    for v in values:
        stripped = v.strip()
        if not stripped:
            continue

        # percentage
        if stripped.endswith("%"):
            scores["percentage"] += 1
            continue

        # currency symbols
        if stripped[0] in "$€£₹" or stripped[:3] in ("USD", "EUR", "GBP", "INR"):
            scores["currency"] += 1
            continue

        # date patterns
        if re.match(r"^\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}$", stripped):
            scores["date"] += 1
            continue
        if re.match(r"^\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2}$", stripped):
            scores["date"] += 1
            continue

        # numeric
        cleaned = stripped.replace(",", "").replace(" ", "")
        try:
            float(cleaned)
            if "." in cleaned:
                scores["decimal"] += 1
            else:
                scores["number"] += 1
            continue
        except ValueError:
            pass

        scores["text"] += 1

    best = max(scores, key=scores.get)
    if scores[best] == 0:
        return "text"
    return best


# ===================================================================
# STEP 2 — Rule-based enterprise defaults (always applied)
# ===================================================================
def apply_rule_based_defaults(filepath: str, structure: list[dict]) -> openpyxl.Workbook:
    """
    Apply baseline professional formatting that guarantees the file
    looks polished even if the LLM step fails.

    Uses the extracted structure to know each sheet's type and header row,
    so it handles title rows, sparse sheets, and hidden/chart sheets correctly.
    """
    # Build a fast lookup: sheet_name → metadata
    sheet_meta = {s["sheet_name"]: s for s in structure}

    wb = openpyxl.load_workbook(filepath)

    thin_side = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for ws in wb.worksheets:
        meta = sheet_meta.get(ws.title)

        # Sheets absent from the structure (hidden, chart) are left untouched
        if meta is None:
            continue

        # Sparse sheets get no structural formatting applied
        if meta.get("sheet_type") == "sparse":
            continue

        if ws.max_row is None or ws.max_column is None or ws.max_row < 1:
            continue

        header_row = meta.get("header_row", 1)
        max_col = min(ws.max_column, MAX_COLS)
        max_row = min(ws.max_row, MAX_FORMAT_ROWS)

        # --- Header row ---
        for c in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.font = Font(name=ENTERPRISE_FONT, bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=thin_side, right=thin_side,
                top=thin_side,
                bottom=Side(style="medium", color="999999"),
            )

        # --- Data cells: borders + font ---
        for r in range(header_row + 1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = Font(name=ENTERPRISE_FONT, size=10)
                cell.alignment = Alignment(vertical="center")

        # --- Auto-fit column widths ---
        for c in range(1, max_col + 1):
            max_len = len(str(ws.cell(row=header_row, column=c).value or ""))
            for r in range(header_row + 1, min(max_row + 1, header_row + 502)):
                cell_len = len(str(ws.cell(row=r, column=c).value or ""))
                max_len = max(max_len, cell_len)
            ws.column_dimensions[get_column_letter(c)].width = max_len + 2

        # --- Row height for header ---
        ws.row_dimensions[header_row].height = 30

        # --- Freeze panes below the header row ---
        ws.freeze_panes = f"A{header_row + 1}"

        # --- Auto-filter anchored at header row ---
        last_col_letter = get_column_letter(max_col)
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{ws.max_row}"

        # --- Print settings: repeat header (and title if present) as print rows ---
        title_row = 1 if meta.get("has_title_row") else header_row
        ws.print_title_rows = f"{title_row}:{header_row}"
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"

    return wb


# ===================================================================
# STEP 3 — LLM one-shot formatting plan
# ===================================================================
def get_llm_formatting_plan(structure: list[dict]) -> dict | None:
    """Send structure to LLM, get back a formatting plan JSON."""
    system_prompt = SYSTEM_PROMPT.format(
        themes=", ".join(THEMES.keys()),
        formats=json.dumps(list(NUMBER_FORMATS.keys())),
        alignments=json.dumps(ALIGNMENTS),
    )
    user_prompt = USER_PROMPT.format(structure=json.dumps(structure, indent=2))

    text = default_llm_service.make_system_user_request(system_prompt, user_prompt)
    if text is None:
        return None

    text = text.strip()
    # Strip markdown code fences if the model includes them anyway
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


# ===================================================================
# STEP 4 — Apply LLM plan onto the workbook
# ===================================================================
def apply_llm_formatting(wb: openpyxl.Workbook, plan: dict) -> openpyxl.Workbook:
    """Layer the LLM-chosen theme and formats on top of rule-based defaults."""

    for sheet_plan in plan.get("sheets", []):
        sheet_name = sheet_plan.get("sheet_name")
        if sheet_name not in wb.sheetnames:
            continue

        # LLM may mark sparse or low-value sheets to skip
        if sheet_plan.get("skip"):
            continue

        ws = wb[sheet_name]
        max_col = min(ws.max_column or 1, MAX_COLS)
        max_row = min(ws.max_row or 1, MAX_FORMAT_ROWS)

        # Header row (LLM echoes back the detected value, may correct if wrong)
        header_row = sheet_plan.get("header_row", 1)

        # Resolve theme
        theme_name = sheet_plan.get("theme", "corporate_blue")
        theme = THEMES.get(theme_name, THEMES["corporate_blue"])

        # --- Sheet tab color ---
        # Preserve the existing color when the LLM judges it is intentional/correct
        if sheet_plan.get("tab_color_action", "override") == "override":
            ws.sheet_properties.tabColor = theme["tab_color"]

        # --- Header styling ---
        header_fill = PatternFill(
            start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid",
        )
        header_font = Font(
            name=ENTERPRISE_FONT, bold=True, size=11, color=theme["header_font"],
        )
        header_border = Border(
            left=Side(style="thin", color=theme["border_dark"]),
            right=Side(style="thin", color=theme["border_dark"]),
            top=Side(style="thin", color=theme["border_dark"]),
            bottom=Side(style="medium", color=theme["border_dark"]),
        )

        for c in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- Alternate row coloring (data rows only, below the header) ---
        if sheet_plan.get("alternate_row_coloring", True) and max_row > header_row + 1:
            alt_fill = PatternFill(
                start_color=theme["accent_light"], end_color=theme["accent_light"], fill_type="solid",
            )
            data_border = Border(
                left=Side(style="thin", color=theme["border_light"]),
                right=Side(style="thin", color=theme["border_light"]),
                top=Side(style="thin", color=theme["border_light"]),
                bottom=Side(style="thin", color=theme["border_light"]),
            )
            for r in range(header_row + 1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = data_border
                    if r % 2 == 0:
                        cell.fill = alt_fill

        # --- Per-column number format + alignment ---
        for col_plan in sheet_plan.get("columns", []):
            col_idx = col_plan.get("index")
            if not col_idx or col_idx > max_col:
                continue

            fmt_key = col_plan.get("number_format", "general")
            fmt_str = NUMBER_FORMATS.get(fmt_key, "General")

            align_key = col_plan.get("alignment", "left")
            if align_key not in ALIGNMENTS:
                align_key = "left"

            for r in range(header_row + 1, max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.number_format = fmt_str
                cell.alignment = Alignment(horizontal=align_key, vertical="center")

        # --- Totals row highlight ---
        if sheet_plan.get("highlight_totals_row") and max_row > header_row + 1:
            total_fill = PatternFill(
                start_color=theme["total_fill"], end_color=theme["total_fill"], fill_type="solid",
            )
            total_font = Font(
                name=ENTERPRISE_FONT, bold=True, size=10, color=theme["total_font"],
            )
            total_border_top = Border(
                top=Side(style="double", color=theme["border_dark"]),
                bottom=Side(style="thin", color=theme["border_dark"]),
                left=Side(style="thin", color=theme["border_light"]),
                right=Side(style="thin", color=theme["border_light"]),
            )
            last_row = ws.max_row  # use actual max, not capped
            for c in range(1, max_col + 1):
                cell = ws.cell(row=last_row, column=c)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border_top

    return wb


# ===================================================================
# MAIN ENTRY POINT
# ===================================================================
def format_excel(input_path: str, output_path: str) -> dict:
    """
    Full pipeline:
      1. Extract structure
      2. Apply rule-based enterprise defaults
      3. Get LLM formatting plan
      4. Apply LLM enhancements
      5. Save formatted workbook

    Returns a status dict with metadata about what was done.
    """
    # Step 1
    structure = extract_structure(input_path)

    # Step 2 — structure passed so defaults use the correct header rows / sheet types
    wb = apply_rule_based_defaults(input_path, structure)

    # Step 3 + 4
    llm_applied = False
    llm_theme = None
    try:
        plan = get_llm_formatting_plan(structure)
        if plan:
            wb = apply_llm_formatting(wb, plan)
            llm_applied = True
            sheets_in_plan = plan.get("sheets", [])
            if sheets_in_plan:
                llm_theme = sheets_in_plan[0].get("theme")
    except Exception as e:
        # LLM failed — rule-based formatting is still in place
        print(f"[excel_formatter] LLM enhancement failed, using rule-based only: {e}")

    # Step 5
    wb.save(output_path)
    wb.close()

    return {
        "sheets_processed": len(structure),
        "total_rows": sum(s.get("total_rows", 0) for s in structure),
        "llm_enhanced": llm_applied,
        "theme_applied": llm_theme,
    }
