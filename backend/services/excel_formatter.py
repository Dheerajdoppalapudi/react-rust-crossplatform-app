import json
import re

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
LLM_MODEL = "gpt-4o"
MAX_COLS = 100          # safety cap
MAX_FORMAT_ROWS = 10000  # cap for styling loops (data is never lost)
ENTERPRISE_FONT = "Calibri"

# ---------------------------------------------------------------------------
# Enterprise color themes — curated for professional reports
# ---------------------------------------------------------------------------
THEMES = {
    "corporate_blue": {
        "header_fill":   "1B3A5C",
        "header_font":   "FFFFFF",
        "accent_light":  "E9EFF5",
        "accent_medium": "C5D5E8",
        "border_dark":   "1B3A5C",
        "border_light":  "B0C4D8",
        "tab_color":     "1B3A5C",
        "total_fill":    "D4E1EE",
        "total_font":    "1B3A5C",
    },
    "executive_gray": {
        "header_fill":   "2D2D2D",
        "header_font":   "FFFFFF",
        "accent_light":  "F2F2F2",
        "accent_medium": "E0E0E0",
        "border_dark":   "2D2D2D",
        "border_light":  "BDBDBD",
        "tab_color":     "2D2D2D",
        "total_fill":    "E0E0E0",
        "total_font":    "2D2D2D",
    },
    "forest_green": {
        "header_fill":   "1B5E20",
        "header_font":   "FFFFFF",
        "accent_light":  "E8F5E9",
        "accent_medium": "C8E6C9",
        "border_dark":   "1B5E20",
        "border_light":  "A5D6A7",
        "tab_color":     "1B5E20",
        "total_fill":    "C8E6C9",
        "total_font":    "1B5E20",
    },
    "burgundy": {
        "header_fill":   "6E1423",
        "header_font":   "FFFFFF",
        "accent_light":  "FBEAEE",
        "accent_medium": "F0C6CF",
        "border_dark":   "6E1423",
        "border_light":  "D4929E",
        "tab_color":     "6E1423",
        "total_fill":    "F0C6CF",
        "total_font":    "6E1423",
    },
    "navy_gold": {
        "header_fill":   "0D1B2A",
        "header_font":   "E8C547",
        "accent_light":  "EFF1F3",
        "accent_medium": "D3D9E0",
        "border_dark":   "0D1B2A",
        "border_light":  "8B9DAF",
        "tab_color":     "0D1B2A",
        "total_fill":    "D3D9E0",
        "total_font":    "0D1B2A",
    },
    "teal_modern": {
        "header_fill":   "00565B",
        "header_font":   "FFFFFF",
        "accent_light":  "E0F2F1",
        "accent_medium": "B2DFDB",
        "border_dark":   "00565B",
        "border_light":  "80CBC4",
        "tab_color":     "00565B",
        "total_fill":    "B2DFDB",
        "total_font":    "00565B",
    },
    "slate_purple": {
        "header_fill":   "4A148C",
        "header_font":   "FFFFFF",
        "accent_light":  "F3E5F5",
        "accent_medium": "E1BEE7",
        "border_dark":   "4A148C",
        "border_light":  "BA68C8",
        "tab_color":     "4A148C",
        "total_fill":    "E1BEE7",
        "total_font":    "4A148C",
    },
}

# ---------------------------------------------------------------------------
# Number format catalogue
# ---------------------------------------------------------------------------
NUMBER_FORMATS = {
    "integer":       "#,##0",
    "decimal_1":     "#,##0.0",
    "decimal_2":     "#,##0.00",
    "decimal_3":     "#,##0.000",
    "percentage":    "0.00%",
    "percentage_0":  "0%",
    "currency_usd":  "$#,##0.00",
    "currency_eur":  "€#,##0.00",
    "currency_gbp":  "£#,##0.00",
    "currency_inr":  "₹#,##0.00",
    "date_ymd":      "YYYY-MM-DD",
    "date_mdy":      "MM/DD/YYYY",
    "date_dmy":      "DD/MM/YYYY",
    "date_long":     "MMMM D, YYYY",
    "time_hm":       "HH:MM",
    "time_hms":      "HH:MM:SS",
    "accounting":    '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)',
    "text":          "@",
    "general":       "General",
    "id":            "0",
}

ALIGNMENTS = ["left", "center", "right"]


# ===================================================================
# STEP 1 — Extract workbook structure (sent to LLM)
# ===================================================================
def extract_structure(filepath: str) -> list[dict]:
    """Read the workbook and return a lightweight summary per sheet."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = []

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row is None or ws.max_column is None or ws.max_row < 1:
            continue

        max_col = min(ws.max_column, MAX_COLS)
        total_rows = ws.max_row

        # --- headers (row 1) ---
        headers = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=1, column=c).value
            headers.append(str(val) if val is not None else "")

        # --- sample rows (up to 5) ---
        sample_rows = []
        for r in range(2, min(total_rows + 1, 7)):
            row_vals = []
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                row_vals.append(str(val) if val is not None else "")
            sample_rows.append(row_vals)

        # --- per-column metadata ---
        columns = []
        for ci in range(max_col):
            samples = [row[ci] for row in sample_rows if row[ci]]
            columns.append({
                "index": ci + 1,
                "letter": get_column_letter(ci + 1),
                "header": headers[ci],
                "sample_values": samples[:4],
                "inferred_type": _infer_type(samples),
                "non_empty_samples": len(samples),
            })

        # --- merged cells ---
        merged = [str(m) for m in ws.merged_cells.ranges][:30]

        # --- detect if last row looks like a totals row ---
        has_totals_row = False
        if total_rows > 2:
            first_cell = str(ws.cell(row=total_rows, column=1).value or "").lower()
            if first_cell in ("total", "totals", "grand total", "sum", "subtotal"):
                has_totals_row = True

        sheets.append({
            "sheet_name": name,
            "total_rows": total_rows,
            "total_columns": max_col,
            "has_totals_row": has_totals_row,
            "merged_cells": merged,
            "columns": columns,
        })

    wb.close()
    return sheets


def _infer_type(values: list[str]) -> str:
    """Heuristic type inference from sample string values."""
    if not values:
        return "text"

    scores = {"number": 0, "decimal": 0, "date": 0, "percentage": 0, "currency": 0, "text": 0}

    for v in values:
        stripped = v.strip()
        if not stripped:
            continue

        # percentage
        if stripped.endswith("%"):
            scores["percentage"] += 1
            continue

        # currency symbols
        if stripped[0] in "$€£₹" or stripped[:3] in ("USD", "EUR", "GBP", "INR"):
            scores["currency"] += 1
            continue

        # date patterns
        if re.match(r"^\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}$", stripped):
            scores["date"] += 1
            continue
        if re.match(r"^\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2}$", stripped):
            scores["date"] += 1
            continue

        # numeric
        cleaned = stripped.replace(",", "").replace(" ", "")
        try:
            float(cleaned)
            if "." in cleaned:
                scores["decimal"] += 1
            else:
                scores["number"] += 1
            continue
        except ValueError:
            pass

        scores["text"] += 1

    best = max(scores, key=scores.get)
    if scores[best] == 0:
        return "text"
    return best

# ===================================================================
# STEP 2 — Rule-based enterprise defaults (always applied)
# ===================================================================
def apply_rule_based_defaults(filepath: str) -> openpyxl.Workbook:
    """
    Apply baseline professional formatting that guarantees the file
    looks polished even if the LLM step fails.
    """
    wb = openpyxl.load_workbook(filepath)

    thin_side = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_column is None or ws.max_row < 1:
            continue

        max_col = min(ws.max_column, MAX_COLS)
        max_row = min(ws.max_row, MAX_FORMAT_ROWS)

        # --- Header row ---
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name=ENTERPRISE_FONT, bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=thin_side, right=thin_side,
                top=thin_side,
                bottom=Side(style="medium", color="999999"),
            )

        # --- Data cells: borders + font ---
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = Font(name=ENTERPRISE_FONT, size=10)
                cell.alignment = Alignment(vertical="center")

        # --- Auto-fit column widths (like double-clicking column border in Excel) ---
        for c in range(1, max_col + 1):
            max_len = len(str(ws.cell(row=1, column=c).value or ""))
            for r in range(2, min(max_row + 1, 502)):  # sample first 500 rows
                cell_len = len(str(ws.cell(row=r, column=c).value or ""))
                max_len = max(max_len, cell_len)
            # Add small padding, no artificial min/max — true auto-fit
            ws.column_dimensions[get_column_letter(c)].width = max_len + 2

        # --- Row height for header ---
        ws.row_dimensions[1].height = 30

        # --- Freeze header row ---
        ws.freeze_panes = "A2"

        # --- Auto-filter ---
        last_col_letter = get_column_letter(max_col)
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        # --- Print settings ---
        ws.print_title_rows = "1:1"
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"

    return wb

# ===================================================================
# STEP 3 — LLM one-shot formatting plan
# ===================================================================
SYSTEM_PROMPT = """\
You are a senior Excel formatting specialist at a Fortune 500 company.
You design clean, professional, enterprise-grade spreadsheet layouts.

Given a workbook structure, produce a JSON formatting plan.

AVAILABLE THEMES (pick one per sheet):
{themes}

AVAILABLE NUMBER FORMATS (pick one per column):
{formats}

AVAILABLE ALIGNMENTS: {alignments}

GUIDELINES:
- Choose a theme that suits the data context (financial → corporate_blue or navy_gold, environmental → forest_green, etc.)
- Different sheets in the same workbook should use the SAME theme for consistency
- Numeric columns → right-aligned. Text → left-aligned. Headers/IDs → center.
- Use the inferred_type to choose the best number_format.
  For "number" use "integer", for "decimal" use "decimal_2",
  for "currency" use the appropriate currency format,
  for "date" use "date_ymd" (unless samples suggest another format),
  for "percentage" use "percentage", for "text" or "id" use "general" or "text".
- Set alternate_row_coloring to true for data-heavy sheets (>10 rows).
- Set highlight_totals_row to true ONLY if has_totals_row is true.
- Set column_group if adjacent columns are logically related (optional).

Respond with ONLY valid JSON — no markdown fences, no explanation.

JSON SCHEMA:
{{
  "sheets": [
    {{
      "sheet_name": "string",
      "theme": "theme_name",
      "alternate_row_coloring": true/false,
      "highlight_totals_row": true/false,
      "columns": [
        {{
          "index": 1,
          "number_format": "format_key",
          "alignment": "left|center|right"
        }}
      ]
    }}
  ]
}}"""

USER_PROMPT = """\
Workbook structure:

{structure}

Generate the formatting plan."""


def get_llm_formatting_plan(structure: list[dict]) -> dict | None:
    """Send structure to LLM, get back a formatting plan JSON."""
    llm = ChatOpenAI(model=LLM_MODEL, temperature=0, request_timeout=30)

    prompt = ChatPromptTemplate.from_messages([
        ("system", SYSTEM_PROMPT),
        ("user", USER_PROMPT),
    ])

    chain = prompt | llm

    response = chain.invoke({
        "themes": ", ".join(THEMES.keys()),
        "formats": json.dumps(list(NUMBER_FORMATS.keys())),
        "alignments": json.dumps(ALIGNMENTS),
        "structure": json.dumps(structure, indent=2),
    })

    text = response.content.strip()
    # Strip markdown code fences if the model includes them anyway
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


# ===================================================================
# STEP 4 — Apply LLM plan onto the workbook
# ===================================================================
def apply_llm_formatting(wb: openpyxl.Workbook, plan: dict) -> openpyxl.Workbook:
    """Layer the LLM-chosen theme and formats on top of rule-based defaults."""

    for sheet_plan in plan.get("sheets", []):
        sheet_name = sheet_plan.get("sheet_name")
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        max_col = min(ws.max_column or 1, MAX_COLS)
        max_row = min(ws.max_row or 1, MAX_FORMAT_ROWS)

        # Resolve theme
        theme_name = sheet_plan.get("theme", "corporate_blue")
        theme = THEMES.get(theme_name, THEMES["corporate_blue"])

        # --- Sheet tab color ---
        ws.sheet_properties.tabColor = theme["tab_color"]

        # --- Header styling ---
        header_fill = PatternFill(
            start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid",
        )
        header_font = Font(
            name=ENTERPRISE_FONT, bold=True, size=11, color=theme["header_font"],
        )
        header_border = Border(
            left=Side(style="thin", color=theme["border_dark"]),
            right=Side(style="thin", color=theme["border_dark"]),
            top=Side(style="thin", color=theme["border_dark"]),
            bottom=Side(style="medium", color=theme["border_dark"]),
        )

        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- Alternate row coloring ---
        if sheet_plan.get("alternate_row_coloring", True) and max_row > 2:
            alt_fill = PatternFill(
                start_color=theme["accent_light"], end_color=theme["accent_light"], fill_type="solid",
            )
            data_border = Border(
                left=Side(style="thin", color=theme["border_light"]),
                right=Side(style="thin", color=theme["border_light"]),
                top=Side(style="thin", color=theme["border_light"]),
                bottom=Side(style="thin", color=theme["border_light"]),
            )
            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = data_border
                    if r % 2 == 0:
                        cell.fill = alt_fill

        # --- Per-column number format + alignment ---
        for col_plan in sheet_plan.get("columns", []):
            col_idx = col_plan.get("index")
            if not col_idx or col_idx > max_col:
                continue

            fmt_key = col_plan.get("number_format", "general")
            fmt_str = NUMBER_FORMATS.get(fmt_key, "General")

            align_key = col_plan.get("alignment", "left")
            if align_key not in ALIGNMENTS:
                align_key = "left"

            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.number_format = fmt_str
                cell.alignment = Alignment(horizontal=align_key, vertical="center")

        # --- Totals row highlight ---
        if sheet_plan.get("highlight_totals_row") and max_row > 2:
            total_fill = PatternFill(
                start_color=theme["total_fill"], end_color=theme["total_fill"], fill_type="solid",
            )
            total_font = Font(
                name=ENTERPRISE_FONT, bold=True, size=10, color=theme["total_font"],
            )
            total_border_top = Border(
                top=Side(style="double", color=theme["border_dark"]),
                bottom=Side(style="thin", color=theme["border_dark"]),
                left=Side(style="thin", color=theme["border_light"]),
                right=Side(style="thin", color=theme["border_light"]),
            )
            last_row = ws.max_row  # use actual max, not capped
            for c in range(1, max_col + 1):
                cell = ws.cell(row=last_row, column=c)
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border_top

    return wb


# ===================================================================
# MAIN ENTRY POINT
# ===================================================================
def format_excel(input_path: str, output_path: str) -> dict:
    """
    Full pipeline:
      1. Extract structure
      2. Apply rule-based enterprise defaults
      3. Get LLM formatting plan
      4. Apply LLM enhancements
      5. Save formatted workbook

    Returns a status dict with metadata about what was done.
    """
    # Step 1
    structure = extract_structure(input_path)

    # Step 2
    wb = apply_rule_based_defaults(input_path)

    # Step 3 + 4
    llm_applied = False
    llm_theme = None
    try:
        plan = get_llm_formatting_plan(structure)
        if plan:
            wb = apply_llm_formatting(wb, plan)
            llm_applied = True
            # Extract theme for reporting
            sheets_in_plan = plan.get("sheets", [])
            if sheets_in_plan:
                llm_theme = sheets_in_plan[0].get("theme")
    except Exception as e:
        # LLM failed — rule-based formatting is still in place
        print(f"[excel_formatter] LLM enhancement failed, using rule-based only: {e}")

    # Step 5
    wb.save(output_path)
    wb.close()

    return {
        "sheets_processed": len(structure),
        "total_rows": sum(s.get("total_rows", 0) for s in structure),
        "llm_enhanced": llm_applied,
        "theme_applied": llm_theme,
    }
